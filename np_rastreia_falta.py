"""
Neste teste vamos adicionar um estágio de data cleanning antes de
trabalhar combinando os melhores algoritmos até aqui utilizados.
Queremos encontrar outliers --> salvando outliers no arquivo.

Após identificar possíveis outliers, queremos encontrar aqueles que geram faltas, isso é, aqueles que derrubam o sistema, todas as variáveis vão para zero.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.stats import zscore

def load_data(filename):
    df = pd.read_csv(filename)
    df_normal = df[df['role'] == 'normal'].copy()
    df_test = df[df['role'] == 'test-0'].copy()
    
    # Remover as colunas "role" e "offset_seconds"
    columns_to_remove = ['role', 'offset_seconds']
    df_normal = df_normal.drop(columns=columns_to_remove)
    df_test = df_test.drop(columns=columns_to_remove)
    
    return df_normal, df_test


def calculate_missing_data(df):
    total_variables = len(df.columns)
    total_missing_data = df.isnull().sum().sum()
    return pd.DataFrame({'Total_Var_Anal': [total_variables], 'Dados_Faltantes': [total_missing_data]})

def find_outliers(df):
    outliers_report = []
    for col in df.columns:
        if df[col].dtype in ['int64', 'float64']:
            z_scores = zscore(df[col])
            outliers_indices = df.index[abs(z_scores) > 3].tolist()
            if outliers_indices:
                outliers_report.append({'Variável': col, 'Índices dos Outliers': outliers_indices, 'Total de Outliers': len(outliers_indices)})
    return pd.DataFrame(outliers_report)


def check_possible_fault(df):
    possible_fault_report = []
    for col in df.columns:
        if df[col].dtype in ['int64', 'float64']:
            outliers_indices = df.index[df[col] != 0].tolist()
            if outliers_indices:
                possible_fault_report.append({'Variável': col, 'Índices dos Registros com Possível Falta': outliers_indices})
    return pd.DataFrame(possible_fault_report)


def save_to_excel(missing_data_normal, missing_data_test, outliers_report_normal, outliers_report_test, possible_fault_report_normal, possible_fault_report_test, output_file):
    # Create a new workbook
    wb = Workbook()

    # Remove the default "Sheet" if it exists
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create worksheets with desired names
    ws_missing_data_normal = wb.create_sheet(title='Normal_Dados_Falt')
    ws_missing_data_test = wb.create_sheet(title='Teste_Dados_Falt')
    ws_outliers_report_normal = wb.create_sheet(title='Normal_Outliers')
    ws_outliers_report_test = wb.create_sheet(title='Test_Outliers')
    ws_possible_fault_normal = wb.create_sheet(title='Normal_Falt_Poss')
    ws_possible_fault_test = wb.create_sheet(title='Test_Poss_Fault')

    # Set column names in worksheets
    missing_data_normal_columns = ['Total_Var_Anal', 'Dados_Faltantes']
    missing_data_test_columns = ['Total_Var_Anal', 'Dados_Faltantes']
    outliers_report_normal_columns = ['Variável', 'Índices dos Outliers', 'Total de Outliers']
    outliers_report_test_columns = ['Variável', 'Índices dos Outliers', 'Total de Outliers']
    possible_fault_report_normal_columns = ['Variável', 'Índices dos Registros com Possível Falta']
    possible_fault_report_test_columns = ['Variável', 'Índices dos Registros com Possível Falta']

    ws_missing_data_normal.append(missing_data_normal_columns)
    ws_missing_data_test.append(missing_data_test_columns)
    ws_outliers_report_normal.append(outliers_report_normal_columns)
    ws_outliers_report_test.append(outliers_report_test_columns)
    ws_possible_fault_normal.append(possible_fault_report_normal_columns)

    # Adding data to worksheets
    for row in dataframe_to_rows(missing_data_normal, index=False, header=True):
        ws_missing_data_normal.append(row)

    for row in dataframe_to_rows(missing_data_test, index=False, header=True):
        ws_missing_data_test.append(row)

    for _, row in outliers_report_normal.iterrows():
        indices = ', '.join(map(str, row['Índices dos Outliers']))
        ws_outliers_report_normal.append([row['Variável'], indices, row['Total de Outliers']])

    for _, row in outliers_report_test.iterrows():
        indices = ', '.join(map(str, row['Índices dos Outliers']))
        ws_outliers_report_test.append([row['Variável'], indices, row['Total de Outliers']])

    for _, row in possible_fault_report_normal.iterrows():
        indices = ', '.join(map(str, row['Índices dos Registros com Possível Falta']))
        ws_possible_fault_normal.append([row['Variável'], indices])

    for _, row in possible_fault_report_test.iterrows():
        indices = ', '.join(map(str, row['Índices dos Registros com Possível Falta']))
        ws_possible_fault_test.append([row['Variável'], indices])

    # Save the workbook
    wb.save(output_file)

# Chamada das funções para obter os dados
df_normal, df_test = load_data('DADOS/Adendo A.2_Conjunto de Dados_DataSet.csv')
missing_data_normal = calculate_missing_data(df_normal)
missing_data_test = calculate_missing_data(df_test)
outliers_report_normal = find_outliers(df_normal)
outliers_report_test = find_outliers(df_test)
possible_fault_report_normal = check_possible_fault(df_normal)
possible_fault_report_test = check_possible_fault(df_test)

# Nome do arquivo de saída
output_file = 'DADOS/POSS_FALT.xlsx'

# Salvar os dados no arquivo Excel
save_to_excel(missing_data_normal, missing_data_test, outliers_report_normal, outliers_report_test, possible_fault_report_normal, possible_fault_report_test, output_file)
