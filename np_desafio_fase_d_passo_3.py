"""
Vamos trabalhar nesta primeira tentativa de solução do problema utilizando o algoritmo RandomForestClassifier

"""


import pandas as pd
from openpyxl import load_workbook

# Carregar a planilha PASSO2.xlsx
workbook = load_workbook('PASSOS/PASSO2.xlsx')

# Acessar as abas de outliers
ws_train_outliers = workbook['Normal_outliers']
ws_test_outliers = workbook['Teste_outliers']

# Remover as linhas duplicadas de cabeçalho
ws_train_outliers.delete_rows(1, 1)
ws_test_outliers.delete_rows(1, 1)

# Adicionar uma linha em branco no topo da planilha
ws_train_outliers.insert_rows(1)
ws_test_outliers.insert_rows(1)

# Nomear as colunas
column_names = ['var_processos', 'outliers',
                'total_outliers', 'val_min', 'val_medio', 'val_max']

for col_index, column_name in enumerate(column_names, start=1):
    ws_train_outliers.cell(row=1, column=col_index, value=column_name)
    ws_test_outliers.cell(row=1, column=col_index, value=column_name)

# Função para converter os valores literais separados por vírgula em uma lista de inteiros


def convert_values(value):
    if value:
        return list(map(int, value.split(', ')))
    return []

# Função para calcular as estatísticas dos outliers


def calculate_statistics(outliers):
    outliers['total_outliers'] = outliers['outliers'].apply(
        lambda x: len(convert_values(x)))
    outliers['val_min'] = outliers['outliers'].apply(
        lambda x: min(convert_values(x)) if x else None)
    outliers['val_medio'] = outliers['outliers'].apply(lambda x: sum(
        convert_values(x)) / len(convert_values(x)) if x else None)
    outliers['val_max'] = outliers['outliers'].apply(
        lambda x: max(convert_values(x)) if x else None)
    return outliers


# Converter as abas em DataFrames
df_train_outliers = pd.DataFrame(ws_train_outliers.values)
df_test_outliers = pd.DataFrame(ws_test_outliers.values)

# Definir o cabeçalho do DataFrame
df_train_outliers.columns = df_train_outliers.iloc[0]
df_test_outliers.columns = df_test_outliers.iloc[0]

# Remover a primeira linha duplicada de cabeçalho
df_train_outliers = df_train_outliers[1:].reset_index(drop=True)
df_test_outliers = df_test_outliers[1:].reset_index(drop=True)

# Calcular as estatísticas para a aba Normal_outliers
df_train_outliers = calculate_statistics(df_train_outliers)

# Calcular as estatísticas para a aba Teste_outliers
df_test_outliers = calculate_statistics(df_test_outliers)

# Converter o DataFrame de volta para uma lista de listas
train_outliers_values = [
    df_train_outliers.columns.tolist()] + df_train_outliers.values.tolist()
test_outliers_values = [
    df_test_outliers.columns.tolist()] + df_test_outliers.values.tolist()

# Atualizar a planilha com os valores calculados
for row_index, row_values in enumerate(train_outliers_values):
    for col_index, value in enumerate(row_values):
        ws_train_outliers.cell(
            row=row_index+2, column=col_index+1, value=value)

for row_index, row_values in enumerate(test_outliers_values):
    for col_index, value in enumerate(row_values):
        ws_test_outliers.cell(row=row_index+2, column=col_index+1, value=value)

# Salvar a planilha atualizada como PASSO3.xlsx
workbook.save('PASSOS/PASSO3.xlsx')
