"""
A partir da separação das bases calcular a estatísticas descrivas e salvar num arquivo excel

Nota: para o cálculo das estatísticas descritivas foram removidas as colunas role e offset_seconds

"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Lendo as folhas 'Normal' e 'Test-0' para dataframes pandas
with pd.ExcelFile('../DADOS/dataset_separado.xlsx') as xls:
    df_normal = pd.read_excel(xls, 'Normal')
    df_test = pd.read_excel(xls, 'Test-0')

# Removendo as colunas 'role' e 'offset_seconds'
df_normal = df_normal.drop(columns=['role', 'offset_seconds'])
df_test = df_test.drop(columns=['role', 'offset_seconds'])

# Calculando as estatísticas descritivas para cada dataframe
desc_normal = df_normal.describe()
desc_test = df_test.describe()

# Criando um novo Workbook
wb = Workbook()

# Removendo a folha padrão criada e adicionando uma nova chamada 'EST_NORMAL'
wb.remove(wb.active)
ws_normal = wb.create_sheet('EST_NORMAL', 0)

# Adicionando estatísticas descritivas à folha 'EST_NORMAL'
for r in dataframe_to_rows(desc_normal, index=True, header=True):
    ws_normal.append(r)

# Adicionando uma nova folha chamada 'EST_TEST_0'
ws_test = wb.create_sheet('EST_TEST_0', 1)

# Adicionando estatísticas descritivas à folha 'EST_TEST_0'
for r in dataframe_to_rows(desc_test, index=True, header=True):
    ws_test.append(r)

# Salvando o workbook como 'RESULT_EST_DESC.xlsx'
wb.save('DADOS/RESULT_EST_DESC.xlsx')
