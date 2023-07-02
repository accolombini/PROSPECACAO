"""
Vamos trabalhar nesta primeira tentativa de solução do problema utilizando o algoritmo RandomForestClassifier

"""

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from scipy.stats import zscore

# Passo 1: Leitura da base e separação em treinamento e teste
df = pd.read_csv('../DADOS/Adendo A.2_Conjunto de Dados_DataSet.csv')

# Converter a coluna "offset_seconds" para formato datetime
df['offset_seconds'] = pd.to_datetime(df['offset_seconds'], unit='s') + pd.DateOffset(years=2023, months=5, days=10)

# Separar a base em treinamento e teste
df_train = df[df['role'] == 'normal'].copy()
df_test = df[df['role'] == 'test-0'].copy()

# Salvar os dados em uma planilha Excel
output_file = '../PASSOS/PASSO1.xlsx'

with pd.ExcelWriter(output_file) as writer:
    df_train.to_excel(writer, sheet_name='Normal', index=False)
    df_test.to_excel(writer, sheet_name='Teste', index=False)

# Passo 2: Detecção de outliers
def detect_outliers(df):
    outliers_report = []
    for col in df.columns:
        if col not in ['role', 'offset_seconds'] and df[col].dtype in ['int64', 'float64']:
            z_scores = zscore(df[col])
            outliers_indices = df.index[abs(z_scores) > 3].tolist()
            if outliers_indices:
                outliers_report.append({'Variável': col, 'Índices dos Outliers': outliers_indices, 'Total de Outliers': len(outliers_indices)})
    return pd.DataFrame(outliers_report)

def save_outliers_to_excel(df_train, df_test, output_file):
    # Criar um novo workbook
    wb = Workbook()

    # Remover a planilha "Sheet" padrão, se existir
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Adicionar planilhas com os nomes desejados
    ws_train_outliers = wb.create_sheet(title='Normal_outliers')
    ws_test_outliers = wb.create_sheet(title='Teste_outliers')

    # Detectar outliers nos conjuntos de treinamento e teste
    train_outliers = detect_outliers(df_train)
    test_outliers = detect_outliers(df_test)

    # Escrever os dados dos outliers nas planilhas
    for _, row in train_outliers.iterrows():
        indices = ', '.join(map(str, row['Índices dos Outliers']))
        ws_train_outliers.append([row['Variável'], indices, row['Total de Outliers']])

    for _, row in test_outliers.iterrows():
        indices = ', '.join(map(str, row['Índices dos Outliers']))
        ws_test_outliers.append([row['Variável'], indices, row['Total de Outliers']])

    # Salvar o workbook
    wb.save(output_file)

# Carregar os dados dos passos anteriores
df_train = pd.read_excel('PASSOS/PASSO1.xlsx', sheet_name='Normal')
df_test = pd.read_excel('PASSOS/PASSO1.xlsx', sheet_name='Teste')
output_file = '../PASSOS/PASSO2.xlsx'

save_outliers_to_excel(df_train, df_test, output_file)


# Passo 3: Análise dos outliers e adição das colunas na planilha
def analyze_outliers(df, outliers_sheet):
    header_row = ['Variável', 'Mínimo', 'Média', 'Máximo']
    outliers_sheet.append(header_row)
    
    for col in df.columns:
        if col not in ['Variável', 'Índices dos Outliers', 'Total de Outliers']:
            outliers_column = [col]
            outliers_min = min(df[col]) if len(df[col]) > 0 else ''
            outliers_mean = df[col].mean() if len(df[col]) > 0 else ''
            outliers_max = max(df[col]) if len(df[col]) > 0 else ''
            outliers_column.extend([outliers_min, outliers_mean, outliers_max])
            outliers_sheet.append(outliers_column)

# Carregar a planilha PASSO2.xlsx
workbook = load_workbook('../PASSOS/PASSO2.xlsx')

# Acessar as abas de outliers
ws_train_outliers = workbook['Normal_outliers']
ws_test_outliers = workbook['Teste_outliers']

# Realizar a análise dos outliers para cada conjunto
analyze_outliers(df_train, ws_train_outliers)
analyze_outliers(df_test, ws_test_outliers)

# Salvar a planilha atualizada
workbook.save('PASSOS/PASSO2.xlsx')
