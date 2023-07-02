"""
Separando o arquivo em Normal e test-0

"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Ler o arquivo CSV
df = pd.read_csv('../DADOS/Adendo A.2_Conjunto de Dados_DataSet.csv')

# Contar registros para cada valor na coluna "role"
count_normal = df[df['role'] == 'normal'].shape[0]
count_teste = df[df['role'] == 'test-0'].shape[0]

# Imprimir o número de registros para cada valor na coluna "role"
print(f"Número de registros com role 'normal': {count_normal}")
print(f"Número de registros com role 'test-0': {count_teste}")

# Dividir os dados com base na coluna "role"
normal_data = df[df['role'] == 'normal']
teste_data = df[df['role'] == 'test-0']

# Criar um objeto Workbook
workbook = Workbook()

# Criar planilhas no arquivo xlsx
worksheet_normal = workbook.create_sheet(title='Normal')
for row in dataframe_to_rows(normal_data, index=False, header=True):
    worksheet_normal.append(row)

worksheet_teste = workbook.create_sheet(title='Teste-0')
for row in dataframe_to_rows(teste_data, index=False, header=True):
    worksheet_teste.append(row)

# Remover a planilha de exemplo criada por padrão
workbook.remove(workbook['Sheet'])

# Salvar o arquivo xlsx
workbook.save('DADOS/saida.xlsx')
