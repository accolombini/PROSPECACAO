"""
Neste teste vamos adicionar um estágio de data cleanning antes de
trabalhar combinando os melhores algoritmos até aqui utilizados.
Queremos encontrar outliers --> salvando outliers no arquivo.

Para identificar os outliers usarei zscore
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.stats import zscore

def load_data(filename):
    df = pd.read_csv(filename)
    df_normal = df[df['role'] == 'normal'].copy()
    df_test = df[df['role'] == 'test-0'].copy()
    return df_normal, df_test

def calculate_missing_data(df):
    total_variables = len(df.columns)
    total_missing_data = df.isnull().sum().sum()
    return pd.DataFrame({'Total de Variáveis': [total_variables], 'Dados Faltantes': [total_missing_data]})

def find_outliers(df):
    outliers_report = []
    for col in df.columns: 
        if df[col].dtype in ['int64', 'float64']:  # Check if the column is numeric
            z_scores = zscore(df[col])
            outliers_indices = df.index[abs(z_scores) > 3].tolist()
            if outliers_indices:
                outliers_report.append({'Variável': col, 'Índices dos Outliers': outliers_indices, 'Total de Outliers': len(outliers_indices)})
    return pd.DataFrame(outliers_report)

def save_to_excel(report_normal, report_test, filename):
    wb = Workbook()

    wb.remove(wb.active)

    ws_normal = wb.create_sheet('RELAT_NORMAL')
    for r in dataframe_to_rows(report_normal, index=False, header=True):
        ws_normal.append(r)

    ws_test = wb.create_sheet('RELAT_TEST')
    for r in dataframe_to_rows(report_test, index=False, header=True):
        ws_test.append(r)

    wb.save(filename)

def format_outliers_report(df):
    df['Índices dos Outliers'] = df['Índices dos Outliers'].apply(lambda x: ', '.join(map(str, x)))
    return df

if __name__ == "__main__":
    df_normal, df_test = load_data('../DADOS/Adendo A.2_Conjunto de Dados_DataSet.csv')

    # After you split your data into df_normal and df_test, exclude 'role' and 'offset_seconds' columns:
    df_normal = df_normal.drop(columns=['role', 'offset_seconds'])
    df_test = df_test.drop(columns=['role', 'offset_seconds'])

    missing_data_normal = calculate_missing_data(df_normal)
    outliers_report_normal = find_outliers(df_normal)
    outliers_report_normal = format_outliers_report(outliers_report_normal)

    missing_data_test = calculate_missing_data(df_test)
    outliers_report_test = find_outliers(df_test)
    outliers_report_test = format_outliers_report(outliers_report_test)

    report_normal = pd.concat([missing_data_normal, outliers_report_normal], 
                              keys=['Dados Faltantes', 'Outliers'])
    
    report_test = pd.concat([missing_data_test, outliers_report_test], 
                            keys=['Dados Faltantes', 'Outliers'])

    save_to_excel(report_normal, report_test, 'DADOS/DATA_FALT_OUT.xlsx')
